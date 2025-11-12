"""
Tax Reporting System untuk SANGKURIANG
Modul pelaporan pajak untuk transaksi pendanaan kripto di Indonesia
"""

import json
import csv
import uuid
from datetime import datetime, timedelta
from enum import Enum
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, asdict
from pathlib import Path
import asyncio
import logging
from decimal import Decimal, ROUND_HALF_UP
import xml.etree.ElementTree as ET
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class TaxType(Enum):
    """Jenis pajak yang berlaku"""
    PPH_21 = "pph_21"  # Pajak penghasilan pasal 21
    PPH_22 = "pph_22"  # Pajak penghasilan pasal 22
    PPH_23 = "pph_23"  # Pajak penghasilan pasal 23
    PPH_25 = "pph_25"  # Pajak penghasilan pasal 25
    PPH_26 = "pph_26"  # Pajak penghasilan pasal 26
    PPN = "ppn"  # Pajak pertambahan nilai
    PBB = "pbb"  # Pajak bumi dan bangunan
    BPHTB = "bphtb"  # Bea perolehan hak atas tanah dan bangunan
    FINAL_TAX = "final_tax"  # Pajak final untuk crypto
    CAPITAL_GAINS = "capital_gains"  # Pajak keuntungan modal


class TransactionType(Enum):
    """Jenis transaksi"""
    FUNDING_CONTRIBUTION = "funding_contribution"  # Kontribusi pendanaan
    FUNDING_WITHDRAWAL = "funding_withdrawal"  # Penarikan dana
    CRYPTO_PURCHASE = "crypto_purchase"  # Pembelian crypto
    CRYPTO_SALE = "crypto_sale"  # Penjualan crypto
    CRYPTO_TRANSFER = "crypto_transfer"  # Transfer crypto
    STAKING_REWARD = "staking_reward"  # Reward staking
    MINING_REWARD = "mining_reward"  # Reward mining
    AIRDROP = "airdrop"  # Airdrop
    FEE_PAYMENT = "fee_payment"  # Pembayaran fee
    INTEREST_INCOME = "interest_income"  # Bunga/interest


class TaxStatus(Enum):
    """Status pajak"""
    PENDING = "pending"
    CALCULATED = "calculated"
    PAID = "paid"
    REPORTED = "reported"
    OVERDUE = "overdue"
    EXEMPTED = "exempted"


class ReportPeriod(Enum):
    """Periode pelaporan"""
    MONTHLY = "monthly"
    QUARTERLY = "quarterly"
    YEARLY = "yearly"


class TaxPayerType(Enum):
    """Jenis wajib pajak"""
    INDIVIDUAL = "individual"
    CORPORATION = "corporation"
    FOREIGN_INDIVIDUAL = "foreign_individual"
    FOREIGN_CORPORATION = "foreign_corporation"


@dataclass
class TaxRate:
    """Struktur tarif pajak"""
    tax_type: TaxType
    rate: Decimal
    minimum_threshold: Decimal
    maximum_threshold: Optional[Decimal]
    description: str
    effective_date: datetime
    expiry_date: Optional[datetime]


@dataclass
class TransactionRecord:
    """Record transaksi untuk pajak"""
    transaction_id: str
    user_id: str
    taxpayer_id: str
    taxpayer_type: TaxPayerType
    transaction_type: TransactionType
    amount_idr: Decimal
    amount_crypto: Optional[Decimal]
    crypto_currency: Optional[str]
    transaction_date: datetime
    exchange_rate: Decimal  # IDR per USD atau kripto
    fees_idr: Decimal
    net_amount_idr: Decimal
    tax_withheld: Decimal
    description: str
    reference_number: str
    counterparty_id: Optional[str]
    location: str  # Lokasi transaksi
    payment_method: str


@dataclass
class TaxLiability:
    """Kewajiban pajak"""
    liability_id: str
    taxpayer_id: str
    tax_type: TaxType
    tax_period: str  # YYYY-MM format
    taxable_amount: Decimal
    tax_rate: Decimal
    tax_amount: Decimal
    tax_credit: Decimal
    net_tax_payable: Decimal
    calculation_date: datetime
    due_date: datetime
    status: TaxStatus
    payment_date: Optional[datetime]
    payment_reference: Optional[str]
    notes: str


@dataclass
class TaxReport:
    """Laporan pajak"""
    report_id: str
    taxpayer_id: str
    tax_period: str
    report_type: ReportPeriod
    generated_date: datetime
    submission_deadline: datetime
    total_transactions: int
    total_taxable_amount: Decimal
    total_tax_liability: Decimal
    total_tax_paid: Decimal
    total_tax_credit: Decimal
    net_tax_payable: Decimal
    status: TaxStatus
    submitted_date: Optional[datetime]
    submitted_by: Optional[str]
    validation_errors: List[str]


@dataclass
class TaxPayment:
    """Pembayaran pajak"""
    payment_id: str
    taxpayer_id: str
    liability_id: str
    payment_amount: Decimal
    payment_date: datetime
    payment_method: str
    bank_reference: str
    tax_type: TaxType
    tax_period: str
    status: TaxStatus
    confirmation_date: Optional[datetime]
    notes: str


class TaxRateDatabase:
    """Database tarif pajak Indonesia"""
    
    def __init__(self):
        self.tax_rates = self._initialize_tax_rates()
        self.special_crypto_rates = self._initialize_crypto_rates()
    
    def _initialize_tax_rates(self) -> Dict[TaxType, List[TaxRate]]:
        """Inisialisasi tarif pajak umum"""
        return {
            TaxType.PPH_21: [
                TaxRate(
                    tax_type=TaxType.PPH_21,
                    rate=Decimal("0.05"),  # 5%
                    minimum_threshold=Decimal("50000000"),  # 50 juta
                    maximum_threshold=Decimal("250000000"),  # 250 juta
                    description="PPh Pasal 21 untuk penghasilan sampai 250 juta",
                    effective_date=datetime(2024, 1, 1),
                    expiry_date=None
                ),
                TaxRate(
                    tax_type=TaxType.PPH_21,
                    rate=Decimal("0.15"),  # 15%
                    minimum_threshold=Decimal("250000000"),  # 250 juta
                    maximum_threshold=Decimal("500000000"),  # 500 juta
                    description="PPh Pasal 21 untuk penghasilan 250-500 juta",
                    effective_date=datetime(2024, 1, 1),
                    expiry_date=None
                ),
                TaxRate(
                    tax_type=TaxType.PPH_21,
                    rate=Decimal("0.25"),  # 25%
                    minimum_threshold=Decimal("500000000"),  # 500 juta
                    maximum_threshold=None,
                    description="PPh Pasal 21 untuk penghasilan di atas 500 juta",
                    effective_date=datetime(2024, 1, 1),
                    expiry_date=None
                )
            ],
            TaxType.PPN: [
                TaxRate(
                    tax_type=TaxType.PPN,
                    rate=Decimal("0.11"),  # 11%
                    minimum_threshold=Decimal("0"),
                    maximum_threshold=None,
                    description="PPN umum",
                    effective_date=datetime(2022, 4, 1),
                    expiry_date=None
                )
            ],
            TaxType.FINAL_TAX: [
                TaxRate(
                    tax_type=TaxType.FINAL_TAX,
                    rate=Decimal("0.10"),  # 10% final
                    minimum_threshold=Decimal("0"),
                    maximum_threshold=None,
                    description="Pajak final untuk keuntungan crypto",
                    effective_date=datetime(2024, 1, 1),
                    expiry_date=None
                )
            ],
            TaxType.CAPITAL_GAINS: [
                TaxRate(
                    tax_type=TaxType.CAPITAL_GAINS,
                    rate=Decimal("0.20"),  # 20%
                    minimum_threshold=Decimal("0"),
                    maximum_threshold=None,
                    description="Pajak keuntungan modal",
                    effective_date=datetime(2024, 1, 1),
                    expiry_date=None
                )
            ]
        }
    
    def _initialize_crypto_rates(self) -> Dict[str, Decimal]:
        """Inisialisasi tarif khusus untuk kripto"""
        return {
            "trading_profit": Decimal("0.10"),  # 10% untuk keuntungan trading
            "mining_income": Decimal("0.25"),  # 25% untuk penghasilan mining
            "staking_reward": Decimal("0.15"),  # 15% untuk reward staking
            "airdrop_income": Decimal("0.25"),  # 25% untuk airdrop
            "defi_yield": Decimal("0.15"),  # 15% untuk yield DeFi
        }
    
    def get_applicable_rate(self, tax_type: TaxType, amount: Decimal, taxpayer_type: TaxPayerType) -> Optional[TaxRate]:
        """Dapatkan tarif pajak yang berlaku"""
        if tax_type not in self.tax_rates:
            return None
        
        applicable_rates = self.tax_rates[tax_type]
        
        for rate in applicable_rates:
            if rate.minimum_threshold <= amount:
                if rate.maximum_threshold is None or amount < rate.maximum_threshold:
                    return rate
        
        return applicable_rates[-1] if applicable_rates else None
    
    def get_crypto_tax_rate(self, crypto_activity: str) -> Decimal:
        """Dapatkan tarif pajak untuk aktivitas kripto"""
        return self.special_crypto_rates.get(crypto_activity, Decimal("0.10"))


class TransactionProcessor:
    """Proses transaksi untuk keperluan pajak"""
    
    def __init__(self, tax_rate_db: TaxRateDatabase):
        self.tax_rate_db = tax_rate_db
        self.transaction_records = {}
        self.exchange_rates = self._initialize_exchange_rates()
    
    def _initialize_exchange_rates(self) -> Dict[str, Decimal]:
        """Inisialisasi kurs tukar"""
        return {
            "USD_IDR": Decimal("15500"),  # 1 USD = 15,500 IDR
            "BTC_IDR": Decimal("950000000"),  # 1 BTC = 950 juta IDR
            "ETH_IDR": Decimal("50000000"),  # 1 ETH = 50 juta IDR
            "BNB_IDR": Decimal("7000000"),  # 1 BNB = 7 juta IDR
            "ADA_IDR": Decimal("8000"),  # 1 ADA = 8,000 IDR
            "SOL_IDR": Decimal("2500000"),  # 1 SOL = 2.5 juta IDR
        }
    
    async def process_transaction(self, transaction_data: Dict[str, Any]) -> Dict[str, Any]:
        """Proses transaksi dan hitung pajak"""
        try:
            # Validasi data transaksi
            required_fields = ["user_id", "taxpayer_id", "taxpayer_type", "transaction_type", "amount_idr"]
            for field in required_fields:
                if field not in transaction_data:
                    return {
                        "success": False,
                        "error": f"Missing required field: {field}"
                    }
            
            # Generate transaction ID
            transaction_id = str(uuid.uuid4())
            
            # Konversi amount crypto ke IDR jika diperlukan
            amount_idr = Decimal(str(transaction_data["amount_idr"]))
            amount_crypto = None
            crypto_currency = transaction_data.get("crypto_currency")
            
            if crypto_currency and "amount_crypto" in transaction_data:
                amount_crypto = Decimal(str(transaction_data["amount_crypto"]))
                # Konversi ke IDR berdasarkan kurs
                exchange_rate_key = f"{crypto_currency.upper()}_IDR"
                if exchange_rate_key in self.exchange_rates:
                    amount_idr = amount_crypto * self.exchange_rates[exchange_rate_key]
            
            # Hitung fee
            fees_idr = Decimal(str(transaction_data.get("fees_idr", "0")))
            
            # Hitung net amount
            net_amount_idr = amount_idr - fees_idr
            
            # Tentukan pajak yang harus dipotong
            tax_withheld = await self._calculate_withholding_tax(transaction_data["transaction_type"], net_amount_idr)
            
            # Buat record transaksi
            transaction_record = TransactionRecord(
                transaction_id=transaction_id,
                user_id=transaction_data["user_id"],
                taxpayer_id=transaction_data["taxpayer_id"],
                taxpayer_type=TaxPayerType(transaction_data["taxpayer_type"]),
                transaction_type=TransactionType(transaction_data["transaction_type"]),
                amount_idr=amount_idr,
                amount_crypto=amount_crypto,
                crypto_currency=crypto_currency,
                transaction_date=datetime.fromisoformat(transaction_data["transaction_date"]),
                exchange_rate=Decimal(str(transaction_data.get("exchange_rate", "1"))),
                fees_idr=fees_idr,
                net_amount_idr=net_amount_idr,
                tax_withheld=tax_withheld,
                description=transaction_data.get("description", ""),
                reference_number=transaction_data.get("reference_number", ""),
                counterparty_id=transaction_data.get("counterparty_id"),
                location=transaction_data.get("location", "Indonesia"),
                payment_method=transaction_data.get("payment_method", "bank_transfer")
            )
            
            self.transaction_records[transaction_id] = transaction_record
            
            return {
                "success": True,
                "transaction_id": transaction_id,
                "tax_withheld": str(tax_withheld),
                "net_amount": str(net_amount_idr),
                "exchange_rate_used": str(transaction_record.exchange_rate)
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    async def _calculate_withholding_tax(self, transaction_type: str, amount: Decimal) -> Decimal:
        """Hitung pajak yang dipotong untuk transaksi"""
        # Logika untuk menentukan apakah perlu pemotongan pajak
        if transaction_type in ["funding_contribution", "crypto_sale", "staking_reward", "mining_reward"]:
            # Gunakan tarif final untuk crypto
            tax_rate = self.tax_rate_db.get_crypto_tax_rate("trading_profit")
            return (amount * tax_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        
        return Decimal("0")
    
    def get_transaction_summary(self, taxpayer_id: str, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """Dapatkan ringkasan transaksi untuk periode tertentu"""
        print(f"DEBUG: Getting transaction summary for {taxpayer_id} from {start_date} to {end_date}")
        print(f"DEBUG: Total transaction records: {len(self.transaction_records)}")
        
        filtered_transactions = [
            txn for txn in self.transaction_records.values()
            if txn.taxpayer_id == taxpayer_id and start_date <= txn.transaction_date <= end_date
        ]
        
        print(f"DEBUG: Filtered transactions: {len(filtered_transactions)}")
        for txn in filtered_transactions:
            print(f"DEBUG: Transaction {txn.transaction_id}: {txn.transaction_type.value}, net_amount: {txn.net_amount_idr}")
        
        # Kelompokkan berdasarkan jenis transaksi
        by_type = {}
        total_tax_withheld = Decimal("0")
        total_amount = Decimal("0")
        
        for txn in filtered_transactions:
            txn_type = txn.transaction_type.value
            if txn_type not in by_type:
                by_type[txn_type] = {
                    "count": 0,
                    "total_amount": Decimal("0"),
                    "total_tax": Decimal("0")
                }
            
            by_type[txn_type]["count"] += 1
            by_type[txn_type]["total_amount"] += txn.net_amount_idr
            by_type[txn_type]["total_tax"] += txn.tax_withheld
            
            total_tax_withheld += txn.tax_withheld
            total_amount += txn.net_amount_idr
        
        return {
            "total_transactions": len(filtered_transactions),
            "total_amount": str(total_amount),
            "total_tax_withheld": str(total_tax_withheld),
            "by_type": {k: {kk: str(vv) for kk, vv in v.items()} for k, v in by_type.items()},
            "transaction_ids": [txn.transaction_id for txn in filtered_transactions]
        }


class TaxCalculator:
    """Kalkulator pajak untuk berbagai jenis transaksi"""
    
    def __init__(self, tax_rate_db: TaxRateDatabase):
        self.tax_rate_db = tax_rate_db
        self.tax_liabilities = {}
    
    async def calculate_tax_liability(self, taxpayer_id: str, tax_period: str, transaction_summary: Dict[str, Any]) -> Dict[str, Any]:
        """Hitung kewajiban pajak untuk periode tertentu"""
        print(f"DEBUG: calculate_tax_liability called with taxpayer_id={taxpayer_id}, tax_period={tax_period}")
        print(f"DEBUG: transaction_summary = {transaction_summary}")
        try:
            # Parse periode pajak (format: YYYY-MM)
            period_date = datetime.strptime(tax_period, "%Y-%m")
            
            # Hitung untuk setiap jenis pajak
            tax_calculations = {}
            total_tax_payable = Decimal("0")
            
            # Pajak final untuk crypto trading
            crypto_trading_income = Decimal(transaction_summary.get("by_type", {}).get("crypto_sale", {}).get("total_amount", "0"))
            print(f"DEBUG: crypto_trading_income = {crypto_trading_income}")
            print(f"DEBUG: by_type crypto_sale = {transaction_summary.get('by_type', {}).get('crypto_sale', {})}")
            if crypto_trading_income > 0:
                final_tax_rate = self.tax_rate_db.get_crypto_tax_rate("trading_profit")
                final_tax = (crypto_trading_income * final_tax_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                
                tax_calculations[TaxType.FINAL_TAX] = {
                    "taxable_amount": str(crypto_trading_income),
                    "tax_rate": str(final_tax_rate),
                    "tax_amount": str(final_tax),
                    "tax_credit": "0",
                    "net_tax_payable": str(final_tax)
                }
                
                total_tax_payable += final_tax
            
            # PPh 21 untuk penghasilan lainnya
            other_income = Decimal("0")
            for income_type in ["staking_reward", "mining_reward", "interest_income"]:
                income = Decimal(transaction_summary.get("by_type", {}).get(income_type, {}).get("total_amount", "0"))
                other_income += income
            
            if other_income > 0:
                # Gunakan tarif progresif untuk PPh 21
                applicable_rate = self.tax_rate_db.get_applicable_rate(TaxType.PPH_21, other_income, TaxPayerType.INDIVIDUAL)
                if applicable_rate:
                    pph21_tax = (other_income * applicable_rate.rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    
                    tax_calculations[TaxType.PPH_21] = {
                        "taxable_amount": str(other_income),
                        "tax_rate": str(applicable_rate.rate),
                        "tax_amount": str(pph21_tax),
                        "tax_credit": "0",
                        "net_tax_payable": str(pph21_tax)
                    }
                    
                    total_tax_payable += pph21_tax
            
            # Generate liability ID
            liability_id = f"TAX-{taxpayer_id}-{tax_period}"
            
            # Hitung tanggal jatuh tempo (contoh: 15 hari setelah akhir bulan)
            due_date = (period_date.replace(day=1) + timedelta(days=32)).replace(day=15)
            
            # Buat record kewajiban pajak
            tax_liability = TaxLiability(
                liability_id=liability_id,
                taxpayer_id=taxpayer_id,
                tax_type=TaxType.FINAL_TAX,  # Default, akan diupdate berdasarkan perhitungan
                tax_period=tax_period,
                taxable_amount=crypto_trading_income + other_income,
                tax_rate=Decimal("0.10"),  # Default rate
                tax_amount=total_tax_payable,
                tax_credit=Decimal("0"),
                net_tax_payable=total_tax_payable,
                calculation_date=datetime.now(),
                due_date=due_date,
                status=TaxStatus.CALCULATED,
                payment_date=None,
                payment_reference=None,
                notes=f"Tax calculation for period {tax_period}"
            )
            
            self.tax_liabilities[liability_id] = tax_liability
            
            return {
                "success": True,
                "liability_id": liability_id,
                "tax_period": tax_period,
                "total_tax_payable": str(total_tax_payable),
                "tax_breakdown": tax_calculations,
                "due_date": due_date.isoformat()
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def get_tax_summary(self, taxpayer_id: str) -> Dict[str, Any]:
        """Dapatkan ringkasan kewajiban pajak"""
        taxpayer_liabilities = [
            liability for liability in self.tax_liabilities.values()
            if liability.taxpayer_id == taxpayer_id
        ]
        
        total_outstanding = Decimal("0")
        total_paid = Decimal("0")
        by_status = {}
        
        for liability in taxpayer_liabilities:
            if liability.status == TaxStatus.PAID:
                total_paid += liability.net_tax_payable
            else:
                total_outstanding += liability.net_tax_payable
            
            status = liability.status.value
            by_status[status] = by_status.get(status, Decimal("0")) + liability.net_tax_payable
        
        return {
            "total_liabilities": len(taxpayer_liabilities),
            "total_outstanding": str(total_outstanding),
            "total_paid": str(total_paid),
            "by_status": {k: str(v) for k, v in by_status.items()},
            "liability_ids": [liability.liability_id for liability in taxpayer_liabilities]
        }


class TaxReportGenerator:
    """Generator laporan pajak dalam berbagai format"""
    
    def __init__(self):
        self.reports = {}
    
    async def generate_tax_report(self, taxpayer_id: str, tax_period: str, report_type: ReportPeriod, 
                                 transaction_summary: Dict[str, Any], tax_calculations: Dict[str, Any]) -> Dict[str, Any]:
        """Generate laporan pajak"""
        try:
            report_id = f"RPT-{taxpayer_id}-{tax_period}-{report_type.value}"
            
            # Hitung total dari summary
            total_transactions = transaction_summary["total_transactions"]
            total_taxable_amount = Decimal(transaction_summary["total_amount"])
            total_tax_liability = Decimal(tax_calculations["total_tax_payable"])
            
            # Tanggal jatuh tempo laporan
            period_date = datetime.strptime(tax_period, "%Y-%m")
            if report_type == ReportPeriod.MONTHLY:
                submission_deadline = (period_date.replace(day=1) + timedelta(days=32)).replace(day=20)
            elif report_type == ReportPeriod.QUARTERLY:
                submission_deadline = (period_date.replace(day=1) + timedelta(days=120)).replace(day=30)
            else:  # YEARLY
                submission_deadline = (period_date.replace(day=1) + timedelta(days=400)).replace(day=31)
            
            # Buat record laporan
            tax_report = TaxReport(
                report_id=report_id,
                taxpayer_id=taxpayer_id,
                tax_period=tax_period,
                report_type=report_type,
                generated_date=datetime.now(),
                submission_deadline=submission_deadline,
                total_transactions=total_transactions,
                total_taxable_amount=total_taxable_amount,
                total_tax_liability=total_tax_liability,
                total_tax_paid=Decimal("0"),  # Akan diupdate saat pembayaran
                total_tax_credit=Decimal("0"),
                net_tax_payable=total_tax_liability,
                status=TaxStatus.CALCULATED,
                submitted_date=None,
                submitted_by=None,
                validation_errors=[]
            )
            
            self.reports[report_id] = tax_report
            
            return {
                "success": True,
                "report_id": report_id,
                "tax_period": tax_period,
                "total_tax_liability": str(total_tax_liability),
                "submission_deadline": submission_deadline.isoformat(),
                "status": "generated"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    async def export_report_to_excel(self, report_id: str, output_path: str) -> Dict[str, Any]:
        """Export laporan ke format Excel"""
        try:
            if not OPENPYXL_AVAILABLE:
                return {
                    "success": False,
                    "error": "openpyxl library not available. Install with: pip install openpyxl"
                }
            
            if report_id not in self.reports:
                return {
                    "success": False,
                    "error": "Report not found"
                }
            
            report = self.reports[report_id]
            
            # Buat workbook Excel
            wb = Workbook()
            
            # Sheet 1: Ringkasan
            ws_summary = wb.active
            ws_summary.title = "Ringkasan Pajak"
            
            # Header
            ws_summary["A1"] = "LAPORAN PAJAK SANGKURIANG"
            ws_summary["A2"] = f"Periode: {report.tax_period}"
            ws_summary["A3"] = f"Generated: {report.generated_date.strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Style header
            header_font = Font(bold=True, size=14)
            ws_summary["A1"].font = header_font
            ws_summary["A2"].font = Font(bold=True)
            ws_summary["A3"].font = Font(bold=True)
            
            # Data ringkasan
            ws_summary["A5"] = "Total Transaksi"
            ws_summary["B5"] = report.total_transactions
            
            ws_summary["A6"] = "Total Jumlah Transaksi (IDR)"
            ws_summary["B6"] = f"Rp {report.total_taxable_amount:,.2f}"
            
            ws_summary["A7"] = "Total Kewajiban Pajak"
            ws_summary["B7"] = f"Rp {report.total_tax_liability:,.2f}"
            
            ws_summary["A8"] = "Pajak yang Telah Dibayar"
            ws_summary["B8"] = f"Rp {report.total_tax_paid:,.2f}"
            
            ws_summary["A9"] = "Sisa Pajak yang Harus Dibayar"
            ws_summary["B9"] = f"Rp {report.net_tax_payable:,.2f}"
            
            ws_summary["A10"] = "Batas Waktu Pelaporan"
            ws_summary["B10"] = report.submission_deadline.strftime('%Y-%m-%d')
            
            # Auto-adjust column widths
            for column in ws_summary.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: Detail Transaksi (placeholder)
            ws_details = wb.create_sheet("Detail Transaksi")
            ws_details["A1"] = "Detail transaksi akan ditampilkan di sini"
            
            # Sheet 3: Kalkulasi Pajak (placeholder)
            ws_calc = wb.create_sheet("Kalkulasi Pajak")
            ws_calc["A1"] = "Kalkulasi pajak akan ditampilkan di sini"
            
            # Simpan file
            wb.save(output_path)
            
            return {
                "success": True,
                "file_path": output_path,
                "format": "excel",
                "sheets": ["Ringkasan Pajak", "Detail Transaksi", "Kalkulasi Pajak"]
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    async def export_report_to_csv(self, report_id: str, output_path: str) -> Dict[str, Any]:
        """Export laporan ke format CSV"""
        try:
            if report_id not in self.reports:
                return {
                    "success": False,
                    "error": "Report not found"
                }
            
            report = self.reports[report_id]
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = [
                    'report_id', 'taxpayer_id', 'tax_period', 'report_type',
                    'total_transactions', 'total_taxable_amount', 'total_tax_liability',
                    'total_tax_paid', 'net_tax_payable', 'status', 'generated_date'
                ]
                
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                
                writer.writerow({
                    'report_id': report.report_id,
                    'taxpayer_id': report.taxpayer_id,
                    'tax_period': report.tax_period,
                    'report_type': report.report_type.value,
                    'total_transactions': report.total_transactions,
                    'total_taxable_amount': str(report.total_taxable_amount),
                    'total_tax_liability': str(report.total_tax_liability),
                    'total_tax_paid': str(report.total_tax_paid),
                    'net_tax_payable': str(report.net_tax_payable),
                    'status': report.status.value,
                    'generated_date': report.generated_date.isoformat()
                })
            
            return {
                "success": True,
                "file_path": output_path,
                "format": "csv"
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    async def export_report_to_xml(self, report_id: str, output_path: str) -> Dict[str, Any]:
        """Export laporan ke format XML (format DJP Online)"""
        try:
            if report_id not in self.reports:
                return {
                    "success": False,
                    "error": "Report not found"
                }
            
            report = self.reports[report_id]
            
            # Buat struktur XML untuk DJP Online
            root = ET.Element("SPT")
            root.set("xmlns", "http://www.pajak.go.id/spt")
            root.set("version", "1.0")
            
            # Header
            header = ET.SubElement(root, "Header")
            ET.SubElement(header, "ReportId").text = report.report_id
            ET.SubElement(header, "TaxPeriod").text = report.tax_period
            ET.SubElement(header, "TaxpayerId").text = report.taxpayer_id
            ET.SubElement(header, "GeneratedDate").text = report.generated_date.isoformat()
            
            # Data Pajak
            tax_data = ET.SubElement(root, "TaxData")
            ET.SubElement(tax_data, "TotalTransactions").text = str(report.total_transactions)
            ET.SubElement(tax_data, "TotalTaxableAmount").text = str(report.total_taxable_amount)
            ET.SubElement(tax_data, "TotalTaxLiability").text = str(report.total_tax_liability)
            ET.SubElement(tax_data, "TotalTaxPaid").text = str(report.total_tax_paid)
            ET.SubElement(tax_data, "NetTaxPayable").text = str(report.net_tax_payable)
            
            # Simpan file XML
            tree = ET.ElementTree(root)
            tree.write(output_path, encoding='utf-8', xml_declaration=True)
            
            return {
                "success": True,
                "file_path": output_path,
                "format": "xml",
                "djp_compatible": True
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }


class TaxPaymentProcessor:
    """Proses pembayaran pajak"""
    
    def __init__(self):
        self.payments = {}
        self.payment_gateways = self._initialize_payment_gateways()
    
    def _initialize_payment_gateways(self) -> Dict[str, Dict[str, Any]]:
        """Inisialisasi payment gateway untuk pembayaran pajak"""
        return {
            "bank_transfer": {
                "name": "Bank Transfer",
                "supported_banks": ["BCA", "Mandiri", "BNI", "BRI"],
                "processing_time": "1-2 business days",
                "fees": Decimal("2500")
            },
            "virtual_account": {
                "name": "Virtual Account",
                "supported_banks": ["BCA", "Mandiri", "BNI", "BRI", "Permata"],
                "processing_time": "real-time",
                "fees": Decimal("1000")
            },
            "e_wallet": {
                "name": "E-Wallet",
                "providers": ["OVO", "GoPay", "DANA", "ShopeePay"],
                "processing_time": "real-time",
                "fees": Decimal("1500")
            }
        }
    
    async def process_tax_payment(self, payment_data: Dict[str, Any]) -> Dict[str, Any]:
        """Proses pembayaran pajak"""
        try:
            # Validasi data pembayaran
            required_fields = ["taxpayer_id", "liability_id", "payment_amount", "payment_method"]
            for field in required_fields:
                if field not in payment_data:
                    return {
                        "success": False,
                        "error": f"Missing required field: {field}"
                    }
            
            # Generate payment ID
            payment_id = f"PAY-{payment_data['taxpayer_id']}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
            
            # Validasi payment method
            payment_method = payment_data["payment_method"]
            if payment_method not in self.payment_gateways:
                return {
                    "success": False,
                    "error": f"Unsupported payment method: {payment_method}"
                }
            
            # Proses pembayaran (simulasi)
            payment_result = await self._simulate_payment_processing(payment_data)
            
            if not payment_result["success"]:
                return payment_result
            
            # Buat record pembayaran
            tax_payment = TaxPayment(
                payment_id=payment_id,
                taxpayer_id=payment_data["taxpayer_id"],
                liability_id=payment_data["liability_id"],
                payment_amount=Decimal(str(payment_data["payment_amount"])),
                payment_date=datetime.now(),
                payment_method=payment_method,
                bank_reference=payment_result["bank_reference"],
                tax_type=TaxType(payment_data.get("tax_type", "final_tax")),
                tax_period=payment_data.get("tax_period", datetime.now().strftime("%Y-%m")),
                status=TaxStatus.PAID,
                confirmation_date=datetime.now(),
                notes=payment_data.get("notes", "")
            )
            
            self.payments[payment_id] = tax_payment
            
            return {
                "success": True,
                "payment_id": payment_id,
                "bank_reference": payment_result["bank_reference"],
                "payment_date": tax_payment.payment_date.isoformat(),
                "processing_fee": str(payment_result["processing_fee"]),
                "total_amount": str(tax_payment.payment_amount + payment_result["processing_fee"])
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    async def _simulate_payment_processing(self, payment_data: Dict[str, Any]) -> Dict[str, Any]:
        """Simulasi proses pembayaran"""
        payment_method = payment_data["payment_method"]
        gateway = self.payment_gateways[payment_method]
        
        # Generate bank reference
        bank_reference = f"REF-{uuid.uuid4().hex[:12].upper()}"
        
        # Hitung biaya pemrosesan
        processing_fee = gateway["fees"]
        
        # Simulasi processing time
        await asyncio.sleep(1)  # Simulasi delay
        
        return {
            "success": True,
            "bank_reference": bank_reference,
            "processing_fee": processing_fee,
            "processing_time": gateway["processing_time"]
        }
    
    def get_payment_history(self, taxpayer_id: str) -> Dict[str, Any]:
        """Dapatkan riwayat pembayaran"""
        taxpayer_payments = [
            payment for payment in self.payments.values()
            if payment.taxpayer_id == taxpayer_id
        ]
        
        total_payments = Decimal("0")
        by_method = {}
        
        for payment in taxpayer_payments:
            total_payments += payment.payment_amount
            method = payment.payment_method
            by_method[method] = by_method.get(method, Decimal("0")) + payment.payment_amount
        
        return {
            "total_payments": len(taxpayer_payments),
            "total_amount": str(total_payments),
            "by_method": {k: str(v) for k, v in by_method.items()},
            "payments": [
                {
                    "payment_id": p.payment_id,
                    "payment_date": p.payment_date.isoformat(),
                    "payment_amount": str(p.payment_amount),
                    "payment_method": p.payment_method,
                    "tax_type": p.tax_type.value,
                    "status": p.status.value
                }
                for p in taxpayer_payments
            ]
        }


class TaxReportingManager:
    """Manajer utama untuk sistem pelaporan pajak"""
    
    def __init__(self, storage_path: str = "tax_reports"):
        self.storage_path = Path(storage_path)
        self.storage_path.mkdir(exist_ok=True)
        
        # Inisialisasi komponen
        self.tax_rate_db = TaxRateDatabase()
        self.transaction_processor = TransactionProcessor(self.tax_rate_db)
        self.tax_calculator = TaxCalculator(self.tax_rate_db)
        self.report_generator = TaxReportGenerator()
        self.payment_processor = TaxPaymentProcessor()
        
        # Setup logging
        self._setup_logging()
    
    def _setup_logging(self):
        """Setup logging untuk audit trail"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(self.storage_path / "tax_reporting.log"),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    async def process_crypto_transaction(self, transaction_data: Dict[str, Any]) -> Dict[str, Any]:
        """Proses transaksi kripto dan hitung pajak"""
        self.logger.info(f"Processing crypto transaction for user {transaction_data.get('user_id', 'unknown')}")
        
        # Proses transaksi
        result = await self.transaction_processor.process_transaction(transaction_data)
        
        if result["success"]:
            self.logger.info(f"Transaction processed successfully: {result['transaction_id']}")
        else:
            self.logger.error(f"Transaction processing failed: {result['error']}")
        
        return result
    
    async def calculate_monthly_tax(self, taxpayer_id: str, year_month: str) -> Dict[str, Any]:
        """Hitung pajak bulanan"""
        print(f"DEBUG: calculate_monthly_tax called with taxpayer_id={taxpayer_id}, year_month={year_month}")
        self.logger.info(f"Calculating monthly tax for taxpayer {taxpayer_id} for period {year_month}")
        
        # Parse periode
        try:
            period_date = datetime.strptime(year_month, "%Y-%m")
            start_date = period_date.replace(day=1)
            end_date = (period_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        except ValueError:
            return {
                "success": False,
                "error": "Invalid period format. Use YYYY-MM format"
            }
        
        # Dapatkan ringkasan transaksi
        print(f"DEBUG: Getting transaction summary for {taxpayer_id} from {start_date} to {end_date}")
        transaction_summary = self.transaction_processor.get_transaction_summary(taxpayer_id, start_date, end_date)
        print(f"DEBUG: Transaction summary: {transaction_summary}")
        
        # Hitung kewajiban pajak
        print(f"DEBUG: Calling calculate_tax_liability with summary: {transaction_summary}")
        tax_calculation = await self.tax_calculator.calculate_tax_liability(taxpayer_id, year_month, transaction_summary)
        print(f"DEBUG: Tax calculation result: {tax_calculation}")
        
        if tax_calculation["success"]:
            self.logger.info(f"Tax calculation completed successfully. Total tax payable: {tax_calculation['total_tax_payable']}")
        else:
            self.logger.error(f"Tax calculation failed: {tax_calculation['error']}")
        
        return tax_calculation
    
    async def generate_tax_report(self, taxpayer_id: str, tax_period: str, report_type: ReportPeriod) -> Dict[str, Any]:
        """Generate laporan pajak"""
        self.logger.info(f"Generating tax report for taxpayer {taxpayer_id} for period {tax_period}")
        
        # Hitung pajak terlebih dahulu
        tax_calculation = await self.calculate_monthly_tax(taxpayer_id, tax_period)
        
        if not tax_calculation["success"]:
            return tax_calculation
        
        # Dapatkan ringkasan transaksi
        try:
            period_date = datetime.strptime(tax_period, "%Y-%m")
            start_date = period_date.replace(day=1)
            end_date = (period_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        except ValueError:
            return {
                "success": False,
                "error": "Invalid period format. Use YYYY-MM format"
            }
        
        transaction_summary = self.transaction_processor.get_transaction_summary(taxpayer_id, start_date, end_date)
        
        # Generate laporan
        report_result = await self.report_generator.generate_tax_report(
            taxpayer_id, tax_period, report_type, transaction_summary, tax_calculation
        )
        
        if report_result["success"]:
            self.logger.info(f"Tax report generated successfully: {report_result['report_id']}")
        else:
            self.logger.error(f"Tax report generation failed: {report_result['error']}")
        
        return report_result
    
    async def export_tax_report(self, report_id: str, format_type: str, output_path: str) -> Dict[str, Any]:
        """Export laporan pajak ke berbagai format"""
        self.logger.info(f"Exporting tax report {report_id} to {format_type} format")
        
        if format_type.lower() == "excel":
            return await self.report_generator.export_report_to_excel(report_id, output_path)
        elif format_type.lower() == "csv":
            return await self.report_generator.export_report_to_csv(report_id, output_path)
        elif format_type.lower() == "xml":
            return await self.report_generator.export_report_to_xml(report_id, output_path)
        else:
            return {
                "success": False,
                "error": f"Unsupported format: {format_type}. Use excel, csv, or xml"
            }
    
    async def process_tax_payment(self, payment_data: Dict[str, Any]) -> Dict[str, Any]:
        """Proses pembayaran pajak"""
        self.logger.info(f"Processing tax payment for taxpayer {payment_data.get('taxpayer_id', 'unknown')}")
        
        result = await self.payment_processor.process_tax_payment(payment_data)
        
        if result["success"]:
            self.logger.info(f"Tax payment processed successfully: {result['payment_id']}")
        else:
            self.logger.error(f"Tax payment processing failed: {result['error']}")
        
        return result
    
    def get_taxpayer_summary(self, taxpayer_id: str) -> Dict[str, Any]:
        """Dapatkan ringkasan status pajak taxpayer"""
        tax_summary = self.tax_calculator.get_tax_summary(taxpayer_id)
        payment_history = self.payment_processor.get_payment_history(taxpayer_id)
        
        return {
            "taxpayer_id": taxpayer_id,
            "tax_summary": tax_summary,
            "payment_history": payment_history,
            "compliance_status": "compliant" if tax_summary["total_outstanding"] == "0" else "pending_payment",
            "last_updated": datetime.now().isoformat()
        }


# Contoh penggunaan
async def main():
    """Demo penggunaan Tax Reporting System"""
    
    # Inisialisasi manajer pelaporan pajak
    tax_manager = TaxReportingManager()
    
    print("=== SANGKURIANG Tax Reporting System Demo ===\n")
    
    # 1. Proses transaksi kripto
    print("1. Processing crypto transaction...")
    transaction_data = {
        "user_id": "USER_001",
        "taxpayer_id": "NPWP_123456789",
        "taxpayer_type": "individual",
        "transaction_type": "crypto_sale",
        "amount_crypto": "0.1",
        "crypto_currency": "BTC",
        "transaction_date": datetime.now().isoformat(),
        "fees_idr": "50000",
        "description": "Sale of 0.1 BTC",
        "reference_number": "TXN-001",
        "location": "Jakarta",
        "payment_method": "bank_transfer"
    }
    
    result = await tax_manager.process_crypto_transaction(transaction_data)
    if result["success"]:
        print(f"   ✓ Transaction processed successfully")
        print(f"   - Transaction ID: {result['transaction_id']}")
        print(f"   - Tax withheld: Rp {result['tax_withheld']}")
    
    print()
    
    # 2. Hitung pajak bulanan
    print("2. Calculating monthly tax...")
    tax_calculation = await tax_manager.calculate_monthly_tax("NPWP_123456789", "2024-01")
    
    if tax_calculation["success"]:
        print(f"   ✓ Tax calculation completed")
        print(f"   - Total tax payable: Rp {tax_calculation['total_tax_payable']}")
        print(f"   - Due date: {tax_calculation['due_date']}")
    
    print()
    
    # 3. Generate laporan pajak
    print("3. Generating tax report...")
    report_result = await tax_manager.generate_tax_report("NPWP_123456789", "2024-01", ReportPeriod.MONTHLY)
    
    if report_result["success"]:
        print(f"   ✓ Tax report generated")
        print(f"   - Report ID: {report_result['report_id']}")
        print(f"   - Status: {report_result['status']}")
    
    print()
    
    # 4. Export laporan ke Excel
    print("4. Exporting tax report to Excel...")
    excel_path = str(tax_manager.storage_path / "tax_report_2024-01.xlsx")
    export_result = await tax_manager.export_tax_report(report_result["report_id"], "excel", excel_path)
    
    if export_result["success"]:
        print(f"   ✓ Report exported to Excel")
        print(f"   - File path: {export_result['file_path']}")
        print(f"   - Sheets: {', '.join(export_result['sheets'])}")
    
    print()
    
    # 5. Proses pembayaran pajak
    print("5. Processing tax payment...")
    payment_data = {
        "taxpayer_id": "NPWP_123456789",
        "liability_id": tax_calculation["liability_id"],
        "payment_amount": tax_calculation["total_tax_payable"],
        "payment_method": "virtual_account",
        "tax_type": "final_tax",
        "tax_period": "2024-01",
        "notes": "Payment for January 2024 crypto taxes"
    }
    
    payment_result = await tax_manager.process_tax_payment(payment_data)
    
    if payment_result["success"]:
        print(f"   ✓ Tax payment processed")
        print(f"   - Payment ID: {payment_result['payment_id']}")
        print(f"   - Bank reference: {payment_result['bank_reference']}")
        print(f"   - Processing fee: Rp {payment_result['processing_fee']}")
    
    print()
    
    # 6. Dapatkan ringkasan taxpayer
    print("6. Getting taxpayer summary...")
    taxpayer_summary = tax_manager.get_taxpayer_summary("NPWP_123456789")
    
    print(f"   ✓ Taxpayer summary retrieved")
    print(f"   - Total outstanding: Rp {taxpayer_summary['tax_summary']['total_outstanding']}")
    print(f"   - Total paid: Rp {taxpayer_summary['tax_summary']['total_paid']}")
    print(f"   - Compliance status: {taxpayer_summary['compliance_status']}")
    
    print("\n=== Demo completed successfully! ===")


if __name__ == "__main__":
    asyncio.run(main())